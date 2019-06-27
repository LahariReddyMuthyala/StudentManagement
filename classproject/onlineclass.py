import click
import openpyxl
import MySQLdb
import os
import sys
from bs4 import BeautifulSoup
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()
from onlineapp.models import *


@click.group()
@click.pass_context
def cli(ctx):
    pass

def connectDB(db_name):
    con = MySQLdb.connect("localhost", "root", "04111998",db_name)
    return con
    pass


def dropdb(ctx,db_name):
    con=connectDB(db_name)
    cur = con.cursor()
    cur.execute("DROP DATABASE IF EXISTS onlineclass")
    con.commit()
    pass

@cli.command()
@click.argument('source_excel1', nargs=1)
@click.argument('source_html', nargs=1)
@click.pass_context
def importdata(ctx,source_excel1, source_html):

    wb1=openpyxl.load_workbook(source_excel1)

    sheet = wb1.get_sheet_by_name('Colleges')
    rows = sheet.max_row
    columns = sheet.max_column
    for r in range(2, rows + 1):
        val = []
        for c in range(1, columns + 1):
            e = sheet.cell(row=r, column=c)
            val.append(e.value)
        valt = tuple(val)
        print(valt)
        c = College(name=valt[0], acronym=valt[1], location=valt[2], contact=valt[3])
        c.save()

    sheet=wb1.get_sheet_by_name('Current')
    rows=sheet.max_row
    columns=sheet.max_column
    for r in range(2,rows+1):
        val=[]
        for c in range(1,columns+1):
            e=sheet.cell(row=r,column=c)
            val.append(e.value)
        valt=tuple(val)
        print(valt)
        s = Student(name = valt[0], college = College.objects.get(acronym = valt[1]), email = valt[2], db_folder = valt[3])
        s.save()

    soup = BeautifulSoup(open(source_html).read(), 'html.parser')
    headings = soup.find_all('th')
    data = soup.find_all('tr')
    for row in range(2, len(data) + 1):
        rowTemp = data[row - 1].find_all('td')
        tds = [rowTemp[i].text for i in range(len(rowTemp))]
        name = tds[1].split('_')
        if(name[2]!="skc"):
            m = MockTest1(student = Student.objects.get(db_folder = name[2]), problem1 = tds[2], problem2 = tds[3], problem3 = tds[4], problem4 = tds[5], total = tds[6] )
            m.save()



if __name__ == '__main__':
    cli(obj={})